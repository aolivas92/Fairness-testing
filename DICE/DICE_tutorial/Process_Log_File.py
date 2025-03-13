import re
import pandas as pd
import os
from collections import defaultdict
from openpyxl import load_workbook

# Define regex patterns for extracting relevant information
instance_pattern = re.compile(r"M_INSTANCE CALL: (\d+)")
tries_pattern = re.compile(r"Finished generating Counter Factual Successfully. Tries: (\d+)")
completion_pattern = re.compile(r"CRITICAL - FAILED TO GENERATE COUNTER FACTUAL, MAX RETRIES HIT")

# Specific error messages
error_types = {
    "FAILED FEATURES.": re.compile(r"ERROR - FAILED TO VERIFY ALL FEATURES."),
    "FAILED SENS. PARAMETER.": re.compile(r"ERROR - FAILED TO CHANGE SENSITIVE PARAMETER."),
    "FAILED ENCODING.": re.compile(r"ERROR - FAILED TO ENCODE SAMPLE: (.+)"),
    "FAILED CLOSEST MATCH": re.compile(r"ERROR - Failed to find a closest match for: (.+)")
}

# Function to process a log file
def process_log_file(log_path, excel_path, file_index):
    data = []
    instance_id = None
    tries = None
    errors = defaultdict(int)  # Dictionary to count each error type
    success = True  # Assume success unless an error indicates otherwise
    all_error_types = set(error_types.keys())  # Predefine all known errors

    with open(log_path, "r") as log_file:
        for line in log_file:
            # Extract test instance ID
            instance_match = instance_pattern.search(line)
            if instance_match:
                if instance_id is not None:  # Save previous instance data before starting a new one
                    data.append([instance_id, tries, sum(errors.values()), success, errors.copy()])

                instance_id = int(instance_match.group(1))
                tries = None
                errors = defaultdict(int)  # Reset error count for new instance
                success = True

            # Extract number of tries for successful cases
            tries_match = tries_pattern.search(line)
            if tries_match:
                tries = int(tries_match.group(1))

            # Extract errors
            for error_name, error_pattern in error_types.items():
                error_match = error_pattern.search(line)
                if error_match:
                    errors[error_name] += 1
                    success = False

            # Check if the test instance failed critically
            if completion_pattern.search(line):
                errors["CRITICAL FAILURE: MAX RETRIES HIT"] += 1
                success = False

        # Append last instance after finishing reading
        if instance_id is not None:
            data.append([instance_id, tries, sum(errors.values()), success, errors.copy()])

    # Create column names
    columns = ["Test Instance", "Tries", "Total Errors", "Success"]
    columns.extend(sorted(all_error_types))
    columns.append("CRITICAL FAILURE: MAX RETRIES HIT")  # Add the critical error column

    # Convert data to structured format
    structured_data = []
    for row in data:
        instance_id, tries, total_errors, success, error_counts = row
        row_data = [instance_id, tries, total_errors, success]

        # Add error counts, ensuring all errors are present
        for error in sorted(all_error_types):
            row_data.append(error_counts.get(error, 0))
        row_data.append(error_counts.get("CRITICAL FAILURE: MAX RETRIES HIT", 0))  # Add critical failure count
        
        structured_data.append(row_data)

    # Convert to DataFrame
    df = pd.DataFrame(structured_data, columns=columns)
    
    # Calculate starting row for this file's data
    # First file starts at row 0, subsequent files leave 2 rows gap from previous data
    start_row = 0
    if file_index > 1 and os.path.exists(excel_path):
        try:
            # Try to get existing workbook to determine last row
            wb = load_workbook(excel_path)
            ws = wb['Log Data']
            last_row = ws.max_row
            # For second file onwards, add a gap of 2 rows plus 1 for the filename header
            start_row = last_row + 3
        except:
            # If there's an issue with the workbook, fallback to a simple calculation
            # Each file after the first starts 3 rows below where it would have in the original script
            # (approximation, as we can't know exact row counts without reading the file)
            start_row = (file_index - 1) * (len(df) + 3)
    
    # Create a one-row DataFrame with the filename
    file_name = log_path.split('/')[6]
    file_header = pd.DataFrame([[log_path]], columns=[file_name])
    
    # Write to Excel file
    if os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            # Write the filename
            file_header.to_excel(writer, sheet_name="Log Data", index=False, startrow=start_row)
            # Write the data starting 1 row below the filename
            df.to_excel(writer, sheet_name="Log Data", index=False, startrow=start_row+1)
    else:
        with pd.ExcelWriter(excel_path, mode="w", engine="openpyxl") as writer:
            # For the first file, write the filename then the data
            file_header.to_excel(writer, sheet_name="Log Data", index=False)
            df.to_excel(writer, sheet_name="Log Data", index=False, startrow=1)

    print(f"Processed {log_path} and saved data to {excel_path}")

# Process all log files
for i in range(1, 11):
    # Specify paths
    log_file_path = f"../results/bank/DICE/RQ1/claude_10max/{i}_10runs_claude/logfile.log"
    excel_file_path = "../results/bank/DICE/RQ1/claude_10max/log_results.xlsx"

    # Process the log file with its index for positioning
    process_log_file(log_file_path, excel_file_path, i)